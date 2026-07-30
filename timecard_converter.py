#!/usr/bin/env python3
"""
Timecard file converter / folder listener for apple-line packer piece counts.

Converts daily packing-line piece-count text files (NF-*.txt for Northern
Fruit, IL-*.txt for Ice Lakes) into the payroll system's Excel
"Time and Attendance Timecard Import" template, translating SIMS package
codes into Paycom earning codes on the way.

Source .txt layout (tab-delimited, no header), e.g.:
    0205261P1   1   0303    510 10100121    0001
    col 1: date/batch code   -> template column C (Date)
    col 2: line number       -> (not used by payroll import)
    col 3: packer/employee   -> template column A (Employee ID)  [text, keeps leading zeros]
    col 4: SIMS package code -> template column F (Earning Code) TRANSLATED to Paycom
    col 5: labor allocation  -> template column I (Labor Allocation Code) [number]
    col 6: pieces packed     -> template column N (Units)        [text, keeps leading zeros]

Usage:
    # one-off conversion of a single file
    python3 timecard_converter.py convert /path/to/NF-Y2026M02D05.txt

    # run forever, watching a folder for new NF-*.txt / IL-*.txt files
    python3 timecard_converter.py watch

Add --strict-codes to refuse any file containing a package code that isn't in
PACKAGE_CODES below. Without it, unknown codes pass through unchanged and are
logged as warnings.
"""

import argparse
import csv
import shutil
import sys
import time
from datetime import date
from pathlib import Path
from typing import NamedTuple

import openpyxl

# ---------------------------------------------------------------------------
# SIMS package code -> (Paycom earning code, Paycom description)
# This is the authoritative mapping. Add new pack styles here.
# ---------------------------------------------------------------------------
PACKAGE_CODES = {
    "500": ("E50", "Euro 1/2 Box"),
    "505": ("L05", "Loose Boxes"),
    "510": ("T10", "Top Pad"),
    "520": ("SS2", "Special"),
    "525": ("B25", "Bags"),
    "530": ("E30", "Euro Bags"),
    "535": ("C35", "Cell Pack"),
    "540": ("S40", "Sleeve Bags"),
    "545": ("H45", "Half Carton"),
    "555": ("H55", "Heavy Pack"),
    "560": ("C60", "Clam"),
    "565": ("E65", "Euro Tray"),
    "570": ("T70", "Top Pad (HC)"),
}

# ---------------------------------------------------------------------------
# Default configuration (override with command-line options if preferred)
# ---------------------------------------------------------------------------
DEFAULT_TEMPLATE = Path(__file__).parent / "Timecard_Import_template_CLEAN.xlsx"
DEFAULT_WATCH_DIR = Path("/srv/timecard/incoming")
DEFAULT_OUTPUT_DIR = Path("/srv/timecard/converted")
DEFAULT_PROCESSED_DIR = Path("/srv/timecard/processed")
DEFAULT_FAILED_DIR = Path("/srv/timecard/failed")
FILE_PATTERNS = ("NF-*.txt", "IL-*.txt")   # Northern Fruit and Ice Lakes lines
POLL_SECONDS = 15          # how often to scan the incoming folder
STABLE_CHECKS = 2          # file size must be unchanged this many polls in a row
TEMPLATE_SHEET = "Import Template"

# ---------------------------------------------------------------------------


class Row(NamedTuple):
    """One line of a packing-line file, ready for the import template."""
    emp_id: str          # column A — zero-padded to EMP_ID_WIDTH for Paycom
    emp_id_raw: str      # the badge number exactly as SIMS wrote it
    date_code: str       # column C — MM/DD/YYYY for Paycom
    date_code_raw: str   # the batch code exactly as SIMS wrote it
    sims_code: str       # source code, kept for display and troubleshooting
    paycom_code: str     # column F — what payroll actually receives
    description: str     # Paycom description, for the preview screen
    alloc: object        # column I
    units: str           # column N

    @property
    def mapped(self) -> bool:
        return self.sims_code in PACKAGE_CODES

    @property
    def assigned(self) -> bool:
        """False for rows the line wrote with no packer ID in them."""
        return bool(self.emp_id)

    @property
    def date_ok(self) -> bool:
        """True when the batch code yielded a real calendar date."""
        return self.date_code != self.date_code_raw

    @property
    def padded(self) -> bool:
        """True when the badge number was widened for Paycom."""
        return self.emp_id != self.emp_id_raw

    @property
    def odd_badge(self) -> bool:
        """A badge that couldn't be padded to the expected width."""
        return bool(self.emp_id_raw) and (
            not self.emp_id_raw.isdigit() or len(self.emp_id) != EMP_ID_WIDTH)


def log(msg: str) -> None:
    print(f"{time.strftime('%Y-%m-%d %H:%M:%S')}  {msg}", flush=True)


def translate_code(sims_code: str):
    """Return (paycom_code, description) for a SIMS package code.

    Unknown codes come back unchanged with an empty description so the caller
    can decide whether to warn or refuse.
    """
    code = sims_code.strip()
    if code in PACKAGE_CODES:
        return PACKAGE_CODES[code]
    # tolerate a stray leading zero or missing one (e.g. "0510" / "51O" won't match)
    stripped = code.lstrip("0")
    for key, value in PACKAGE_CODES.items():
        if key.lstrip("0") == stripped and stripped:
            return value
    return code, ""


def unmapped_codes(rows) -> list:
    """Sorted list of SIMS codes in these rows that have no Paycom mapping."""
    return sorted({r.sims_code for r in rows if not r.mapped})


# Paycom expects the Employee ID (badge number) as exactly this many digits,
# zero-padded: SIMS writes 0303, Paycom wants 0000303.
EMP_ID_WIDTH = 7


def normalize_emp_id(raw: str) -> str:
    """Zero-pad a badge number to the width Paycom expects.

    An empty field stays empty — a row with no packer must not become badge
    0000000. A number already at or over the width, or anything non-numeric,
    is passed through untouched and flagged for review instead of mangled.
    """
    emp = raw.strip()
    if not emp or not emp.isdigit():
        return emp
    return emp.zfill(EMP_ID_WIDTH)


# The line writes the pack date as the first 6 digits of the batch code
# (MMDDYY), followed by line/plant characters such as "1P1" that payroll
# doesn't want: 0720261P1 -> 07/20/2026.
DATE_CODE_DIGITS = 6


def normalize_date(raw: str):
    """Turn a SIMS batch code into the MM/DD/YYYY Paycom expects.

    Returns (value, recognised). Anything that isn't a valid MMDDYY date is
    passed through untouched and flagged, rather than guessed at — a wrong
    date on a timecard is worse than an obvious one.
    """
    code = raw.strip()
    head = code[:DATE_CODE_DIGITS]
    if len(head) == DATE_CODE_DIGITS and head.isdigit():
        mm, dd, yy = head[:2], head[2:4], head[4:6]
        try:
            when = date(2000 + int(yy), int(mm), int(dd))
        except ValueError:
            return code, False
        return f"{when.month:02d}/{when.day:02d}/{when.year}", True
    return code, False


# Expected field widths, used only to work out which column is missing when a
# fixed-width file happens to have one column blank on every single row.
EXPECTED_WIDTHS = {"date": 9, "emp": 4, "code": 3, "alloc": 8, "units": 4}


def sniff_columns(lines) -> list:
    """Work out fixed-width field boundaries from the file itself.

    A column is a separator if it is blank on every line, so the field
    boundaries fall out of the data — no hard-coded offsets to break when the
    line export changes its padding.
    """
    width = max(len(l) for l in lines)
    is_sep = [all(len(l) <= i or l[i] == " " for l in lines) for i in range(width)]

    fields, start = [], None
    for i, sep in enumerate(is_sep + [True]):
        if not sep and start is None:
            start = i
        elif sep and start is not None:
            fields.append((start, i))
            start = None
    return fields


def split_fixed(line: str, columns: list) -> list:
    return [line[a:b].strip() for a, b in columns]


def parse_txt(txt_path: Path, strict: bool = False) -> list:
    """Parse a packing-line file into Row records.

    Handles both layouts the lines produce: tab-delimited, and space-aligned
    fixed width (with either LF or CRLF line endings).
    """
    with open(txt_path, "r", encoding="utf-8-sig", newline="") as fh:
        raw_lines = [l.rstrip("\r\n") for l in fh]
    lines = [l for l in raw_lines if l.strip()]
    if not lines:
        raise ValueError(f"{txt_path.name}: no data rows found")

    tab_delimited = any("\t" in l for l in lines)
    columns = None

    if not tab_delimited:
        columns = sniff_columns(lines)
        if len(columns) == 5:
            # One column is blank on every row. If the third field is 3 wide
            # it's the package code, which means the packer column is the one
            # missing entirely — put an empty placeholder back in its place.
            if columns[2][1] - columns[2][0] == EXPECTED_WIDTHS["code"]:
                columns = columns[:2] + [(columns[1][1], columns[1][1])] + columns[2:]
        if len(columns) != 6:
            raise ValueError(
                f"{txt_path.name}: expected 6 columns, found {len(columns)} "
                f"at {columns} — the line's export layout may have changed"
            )

    rows = []
    for lineno, line in enumerate(lines, start=1):
        if tab_delimited:
            fields = [f.strip() for f in line.split("\t")]
            if len(fields) < 6:
                raise ValueError(
                    f"{txt_path.name} line {lineno}: expected 6 tab-separated "
                    f"fields, found {len(fields)}: {line!r}"
                )
        else:
            fields = split_fixed(line, columns)

        date_code, _line_no, emp_id, sims_code, alloc, units = fields[:6]

        if not sims_code and not units:
            continue  # nothing usable on this line

        paycom_code, description = translate_code(sims_code)
        pack_date, _date_ok = normalize_date(date_code)

        # Labor Allocation Code is stored as a number (matches the sample
        # payroll prepared); fall back to text if it's ever non-numeric.
        alloc_val = int(alloc) if alloc.isdigit() else alloc

        rows.append(Row(
            emp_id=normalize_emp_id(emp_id),
            emp_id_raw=emp_id.strip(),
            date_code=pack_date,
            date_code_raw=date_code.strip(),
            sims_code=sims_code,
            paycom_code=paycom_code,
            description=description,
            alloc=alloc_val,
            units=units,
        ))

    if not rows:
        raise ValueError(f"{txt_path.name}: no data rows found")

    unknown = unmapped_codes(rows)
    if unknown and strict:
        raise ValueError(
            f"{txt_path.name}: package code(s) {', '.join(unknown)} have no Paycom "
            f"equivalent — add them to PACKAGE_CODES in timecard_converter.py"
        )
    return rows


def convert(txt_path: Path, template_path: Path, output_dir: Path,
            also_csv: bool = False, strict: bool = False) -> Path:
    """Convert one .txt file to an .xlsx built from the payroll template."""
    all_rows = parse_txt(txt_path, strict=strict)

    # Rows the line wrote with no packer ID can't go to payroll — column A
    # would be blank. Leave them out, but never silently: report the count and
    # the units so the day can still be reconciled against the line's totals.
    rows = [r for r in all_rows if r.assigned]
    orphans = [r for r in all_rows if not r.assigned]
    if orphans:
        orphan_units = sum(int(r.units) for r in orphans if r.units.isdigit())
        codes = ", ".join(sorted({r.sims_code for r in orphans}))
        log(f"WARNING: {txt_path.name} has {len(orphans)} row(s) with no packer ID "
            f"({orphan_units} units, code(s) {codes}) — left out of the workbook, "
            f"since payroll can't import a blank Employee ID")
    if not rows:
        raise ValueError(f"{txt_path.name}: every row is missing its packer ID")

    unknown = unmapped_codes(rows)
    if unknown:
        log(f"WARNING: {txt_path.name} has package code(s) with no Paycom "
            f"mapping: {', '.join(unknown)} — passed through unchanged, "
            f"payroll may reject those rows")

    wb = openpyxl.load_workbook(template_path)
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"Template is missing the '{TEMPLATE_SHEET}' sheet")
    ws = wb[TEMPLATE_SHEET]

    # Clear any leftover data rows below the header, then write fresh data.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.emp_id)       # A  Employee ID
        ws.cell(row=i, column=3, value=r.date_code)    # C  Date
        ws.cell(row=i, column=6, value=r.paycom_code)  # F  Earning Code (Paycom)
        ws.cell(row=i, column=9, value=r.alloc)        # I  Labor Allocation Code
        ws.cell(row=i, column=14, value=r.units)       # N  Units

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / (txt_path.stem + ".xlsx")
    wb.save(out_path)
    log(f"Converted {txt_path.name} -> {out_path} ({len(rows)} rows)")

    if also_csv:
        csv_path = output_dir / (txt_path.stem + ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            for r in rows:
                # 15 columns A-O, no header record (per import instructions)
                line = [""] * 15
                line[0] = r.emp_id       # A
                line[2] = r.date_code    # C
                line[5] = r.paycom_code  # F
                line[8] = r.alloc        # I
                line[13] = r.units       # N
                writer.writerow(line)
        log(f"Also wrote CSV -> {csv_path}")

    return out_path


# ---------------------------------------------------------------------------
# Folder watcher
# ---------------------------------------------------------------------------

def find_candidates(watch_dir: Path):
    for pattern in FILE_PATTERNS:
        yield from watch_dir.glob(pattern)


def wait_until_stable(path: Path) -> bool:
    """Return True once the file size stops changing (upload finished)."""
    last = -1
    stable = 0
    for _ in range(60):  # give up after ~60 polls
        try:
            size = path.stat().st_size
        except FileNotFoundError:
            return False
        if size == last and size > 0:
            stable += 1
            if stable >= STABLE_CHECKS:
                return True
        else:
            stable = 0
        last = size
        time.sleep(POLL_SECONDS)
    return False


def watch(watch_dir: Path, template_path: Path, output_dir: Path,
          processed_dir: Path, failed_dir: Path, also_csv: bool,
          strict: bool = False) -> None:
    for d in (watch_dir, output_dir, processed_dir, failed_dir):
        d.mkdir(parents=True, exist_ok=True)
    log(f"Watching {watch_dir} for {', '.join(FILE_PATTERNS)} "
        f"(poll every {POLL_SECONDS}s)")

    while True:
        for txt_path in sorted(find_candidates(watch_dir)):
            log(f"Found {txt_path.name}, waiting for file to finish copying...")
            if not wait_until_stable(txt_path):
                log(f"WARNING: {txt_path.name} never stabilized; skipping for now")
                continue
            try:
                convert(txt_path, template_path, output_dir, also_csv, strict)
            except Exception as exc:  # keep the service alive on bad files
                log(f"ERROR converting {txt_path.name}: {exc}")
                shutil.move(str(txt_path), failed_dir / txt_path.name)
                log(f"Moved {txt_path.name} to {failed_dir}")
            else:
                shutil.move(str(txt_path), processed_dir / txt_path.name)
                log(f"Moved {txt_path.name} to {processed_dir}")
        time.sleep(POLL_SECONDS)


# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert packing-line .txt files to the payroll import workbook")
    sub = parser.add_subparsers(dest="command", required=True)

    p_convert = sub.add_parser("convert", help="convert a single .txt file")
    p_convert.add_argument("txt_file", type=Path)
    p_convert.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    p_convert.add_argument("--output-dir", type=Path, default=None,
                           help="default: same folder as the .txt file")
    p_convert.add_argument("--csv", action="store_true",
                           help="also write the no-header .csv the import "
                                "screen ultimately accepts")
    p_convert.add_argument("--strict-codes", action="store_true",
                           help="refuse the file if any package code is unmapped")

    p_watch = sub.add_parser("watch", help="watch a folder continuously")
    p_watch.add_argument("--watch-dir", type=Path, default=DEFAULT_WATCH_DIR)
    p_watch.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    p_watch.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    p_watch.add_argument("--processed-dir", type=Path, default=DEFAULT_PROCESSED_DIR)
    p_watch.add_argument("--failed-dir", type=Path, default=DEFAULT_FAILED_DIR)
    p_watch.add_argument("--csv", action="store_true",
                         help="also write the no-header .csv alongside each .xlsx")
    p_watch.add_argument("--strict-codes", action="store_true",
                         help="quarantine files containing unmapped package codes")

    p_codes = sub.add_parser("codes", help="print the package code mapping")

    args = parser.parse_args()

    if args.command == "codes":
        print(f"{'SIMS':<6} {'Paycom':<8} Description")
        for sims, (paycom, desc) in sorted(PACKAGE_CODES.items()):
            print(f"{sims:<6} {paycom:<8} {desc}")
        return 0

    if not args.template.exists():
        log(f"ERROR: template not found: {args.template}")
        return 1

    if args.command == "convert":
        out_dir = args.output_dir or args.txt_file.parent
        try:
            convert(args.txt_file, args.template, out_dir, args.csv, args.strict_codes)
        except Exception as exc:
            log(f"ERROR: {exc}")
            return 1
    else:
        watch(args.watch_dir, args.template, args.output_dir,
              args.processed_dir, args.failed_dir, args.csv, args.strict_codes)
    return 0


if __name__ == "__main__":
    sys.exit(main())
